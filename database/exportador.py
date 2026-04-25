import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def formatear_moneda(valor: float) -> str:
    """Formatea un número como moneda COP: $5.000"""
    return f"${valor:,.0f}".replace(",", ".")


def exportar_liquidaciones(datos: list[dict], ruta_destino: str | None = None) -> str:
    """
    Genera un archivo .xlsx con las liquidaciones.
    Retorna la ruta del archivo generado.
    """
    if not ruta_destino:
        escritorio = os.path.join(os.path.expanduser("~"), "Escritorio")
        if not os.path.exists(escritorio):
            escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
        if not os.path.exists(escritorio):
            escritorio = os.path.expanduser("~")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        ruta_destino = os.path.join(escritorio, f"Liquidaciones_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Liquidaciones"

    # ── Estilos ──
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_font = Font(name="Calibri", size=10)
    money_font = Font(name="Calibri", size=10, color="2ecc71")
    border = Border(
        left=Side(style="thin", color="333333"),
        right=Side(style="thin", color="333333"),
        top=Side(style="thin", color="333333"),
        bottom=Side(style="thin", color="333333"),
    )
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # ── Título del reporte ──
    ws.merge_cells("A1:L1")
    titulo_cell = ws["A1"]
    titulo_cell.value = "REPORTE DE LIQUIDACIONES — MENSAJERÍA"
    titulo_cell.font = Font(name="Calibri", bold=True, size=14, color="1a1a2e")
    titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:L2")
    fecha_cell = ws["A2"]
    fecha_cell.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    fecha_cell.font = Font(name="Calibri", italic=True, size=9, color="666666")
    fecha_cell.alignment = Alignment(horizontal="center")

    # ── Encabezados ──
    headers = ["ID", "Mensajero", "Teléfono", "Fecha", "N° Servicios", "Subtotal Servicios",
               "Comisión (20%)", "Aseo", "Base Prestada", "Neto Mensajero", "Ganancia Empresa", "Domicilios"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # ── Datos ──
    from database import database as db
    for row_num, liq in enumerate(datos, 5):
        comision = liq.get("comision_empresa", 0)
        aseo = liq.get("descuento_aseo", 0)
        ganancia_empresa = comision + aseo
        servicios_liq = db.obtener_servicios_por_liquidacion(liq["id"])
        descripciones = ", ".join([s.get("descripcion", "") for s in servicios_liq if s.get("descripcion")])

        valores = [
            liq.get("id", ""),
            liq.get("mensajero_nombre", ""),
            liq.get("mensajero_telefono", ""),
            liq.get("fecha", ""),
            liq.get("num_servicios", 0),
            formatear_moneda(liq.get("subtotal_servicios", 0)),
            formatear_moneda(comision),
            formatear_moneda(aseo),
            formatear_moneda(liq.get("base_prestada", 0)),
            formatear_moneda(liq.get("neto_mensajero", 0)),
            formatear_moneda(ganancia_empresa),
            descripciones
        ]
        for col_num, valor in enumerate(valores, 1):
            cell = ws.cell(row=row_num, column=col_num, value=valor)
            cell.font = money_font if 6 <= col_num <= 11 else cell_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if col_num <= 5 else "right")
            if (row_num - 5) % 2 == 1:
                cell.fill = alt_fill

    # ── Totales ──
    if datos:
        fila_total = len(datos) + 5
        ws.cell(row=fila_total, column=4, value="TOTALES:").font = Font(bold=True, size=11)
        total_subtotal = sum(d.get("subtotal_servicios", 0) for d in datos)
        total_comision = sum(d.get("comision_empresa", 0) for d in datos)
        total_aseo = sum(d.get("descuento_aseo", 0) for d in datos)
        total_base = sum(d.get("base_prestada", 0) for d in datos)
        total_neto = sum(d.get("neto_mensajero", 0) for d in datos)
        total_ganancia_empresa = sum(d.get("comision_empresa", 0) + d.get("descuento_aseo", 0) for d in datos)

        ws.cell(row=fila_total, column=6, value=formatear_moneda(total_subtotal)).font = Font(bold=True, color="2ecc71")
        ws.cell(row=fila_total, column=7, value=formatear_moneda(total_comision)).font = Font(bold=True, color="e74c3c")
        ws.cell(row=fila_total, column=8, value=formatear_moneda(total_aseo)).font = Font(bold=True, color="e67e22")
        ws.cell(row=fila_total, column=9, value=formatear_moneda(total_base)).font = Font(bold=True, color="e67e22")
        ws.cell(row=fila_total, column=10, value=formatear_moneda(total_neto)).font = Font(bold=True, color="2ecc71")
        ws.cell(row=fila_total, column=11, value=formatear_moneda(total_ganancia_empresa)).font = Font(bold=True, color="1a1a2e", size=12)

    # ── Ajustar ancho de columnas ──
    anchos = [8, 22, 15, 22, 15, 22, 18, 12, 15, 18, 20, 40]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    wb.save(ruta_destino)
    return ruta_destino


def exportar_servicios_pendientes(ruta_destino: str | None = None) -> str:
    """
    Exporta a Excel todos los servicios pendientes (sin liquidar) de cada mensajero.
    Genera una hoja general y una hoja por mensajero que tenga servicios.
    Retorna la ruta del archivo generado.
    """
    from database import database as db

    if not ruta_destino:
        escritorio = os.path.join(os.path.expanduser("~"), "Escritorio")
        if not os.path.exists(escritorio):
            escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
        if not os.path.exists(escritorio):
            escritorio = os.path.expanduser("~")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        ruta_destino = os.path.join(escritorio, f"Respaldo_Servicios_Pendientes_{timestamp}.xlsx")

    wb = Workbook()

    # ── Estilos base ──
    header_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill   = PatternFill(start_color="16213e", end_color="16213e", fill_type="solid")
    header_align  = Alignment(horizontal="center", vertical="center")
    title_font    = Font(name="Calibri", bold=True, size=14, color="16213e")
    sub_font      = Font(name="Calibri", italic=True, size=9, color="666666")
    cell_font     = Font(name="Calibri", size=10)
    money_font    = Font(name="Calibri", size=10, color="27ae60")
    warn_fill     = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
    alt_fill      = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    mensajeros = db.obtener_mensajeros()
    todos_los_servicios = []  # Para la hoja resumen

    for mensajero in mensajeros:
        pendientes = db.obtener_servicios_pendientes(mensajero["id"])
        if not pendientes:
            continue

        # Acumular para hoja resumen
        for s in pendientes:
            todos_los_servicios.append({
                **s,
                "mensajero_nombre": mensajero["nombre"],
                "mensajero_telefono": mensajero["telefono"],
            })

        # ── Hoja individual por mensajero ──
        nombre_hoja = mensajero["nombre"][:28].strip()  # Excel max 31 chars
        ws = wb.create_sheet(title=nombre_hoja)

        # Título
        ws.merge_cells("A1:F1")
        ws["A1"].value = f"Servicios Pendientes — {mensajero['nombre']}"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:F2")
        ws["A2"].value = (
            f"Tel: {mensajero['telefono']}   |   "
            f"Exportado: {datetime.now().strftime('%d/%m/%Y %H:%M')}   |   "
            f"Total servicios: {len(pendientes)}"
        )
        ws["A2"].font = sub_font
        ws["A2"].alignment = Alignment(horizontal="center")

        # Advertencia
        ws.merge_cells("A3:F3")
        ws["A3"].value = "⚠️  RESPALDO DE SEGURIDAD — Servicios sin liquidar"
        ws["A3"].font = Font(name="Calibri", bold=True, size=10, color="856404")
        ws["A3"].fill = warn_fill
        ws["A3"].alignment = Alignment(horizontal="center")

        # Encabezados
        cols = ["ID Servicio", "Valor", "Descripción / Cliente", "Fecha y Hora", "Base Mensajero", "Días pendiente"]
        for c, h in enumerate(cols, 1):
            cell = ws.cell(row=4, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
        ws.row_dimensions[4].height = 22

        # Datos
        hoy = datetime.now().date()
        total_valor = 0.0
        for i, s in enumerate(pendientes, 5):
            try:
                fecha_dt = datetime.strptime(s["fecha"], "%Y-%m-%d %H:%M:%S")
                dias_pend = (hoy - fecha_dt.date()).days
            except Exception:
                fecha_dt = s["fecha"]
                dias_pend = "?"

            total_valor += s.get("valor", 0)
            fondo = alt_fill if (i - 5) % 2 == 1 else None

            fila = [
                s.get("id", ""),
                formatear_moneda(s.get("valor", 0)),
                s.get("descripcion", "") or "",
                s["fecha"],
                formatear_moneda(mensajero.get("base_actual", 0)) if i == 5 else "",
                dias_pend,
            ]
            for c, v in enumerate(fila, 1):
                cell = ws.cell(row=i, column=c, value=v)
                cell.font = money_font if c == 2 else cell_font
                cell.border = border
                cell.alignment = Alignment(horizontal="center" if c in (1, 2, 4, 5, 6) else "left")
                if fondo:
                    cell.fill = fondo

        # Fila de total
        fila_total = len(pendientes) + 5
        ws.cell(row=fila_total, column=1, value="TOTAL:").font = Font(bold=True, size=11)
        ws.cell(row=fila_total, column=2, value=formatear_moneda(total_valor)).font = Font(bold=True, color="27ae60", size=11)

        # Anchos de columna
        for i, ancho in enumerate([12, 14, 35, 22, 16, 14], 1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

    # ── Hoja resumen general ──
    ws_resumen = wb.active if not wb.worksheets else wb.create_sheet(title="RESUMEN GENERAL", index=0)
    # Si ya existe la hoja por defecto vacía (Sheet), la renombramos
    if wb.worksheets and wb.worksheets[0].title == "Sheet":
        wb.worksheets[0].title = "RESUMEN GENERAL"
        ws_resumen = wb.worksheets[0]

    ws_resumen.merge_cells("A1:H1")
    ws_resumen["A1"].value = "RESPALDO COMPLETO — SERVICIOS PENDIENTES DE LIQUIDAR"
    ws_resumen["A1"].font = title_font
    ws_resumen["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_resumen.row_dimensions[1].height = 30

    ws_resumen.merge_cells("A2:H2")
    ws_resumen["A2"].value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}   |   Total servicios: {len(todos_los_servicios)}"
    ws_resumen["A2"].font = sub_font
    ws_resumen["A2"].alignment = Alignment(horizontal="center")

    ws_resumen.merge_cells("A3:H3")
    ws_resumen["A3"].value = "⚠️  Este archivo es un respaldo de seguridad. Los datos provienen de la base de datos activa."
    ws_resumen["A3"].font = Font(name="Calibri", bold=True, size=10, color="856404")
    ws_resumen["A3"].fill = warn_fill
    ws_resumen["A3"].alignment = Alignment(horizontal="center")

    cols_res = ["ID", "Mensajero", "Teléfono", "Valor", "Descripción / Cliente", "Fecha y Hora", "Días pendiente", "Base Mensajero"]
    for c, h in enumerate(cols_res, 1):
        cell = ws_resumen.cell(row=4, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    ws_resumen.row_dimensions[4].height = 22

    hoy = datetime.now().date()
    total_general = 0.0
    for i, s in enumerate(todos_los_servicios, 5):
        try:
            fecha_dt = datetime.strptime(s["fecha"], "%Y-%m-%d %H:%M:%S")
            dias_pend = (hoy - fecha_dt.date()).days
        except Exception:
            dias_pend = "?"

        total_general += s.get("valor", 0)
        fondo = alt_fill if (i - 5) % 2 == 1 else None

        fila = [
            s.get("id", ""),
            s.get("mensajero_nombre", ""),
            s.get("mensajero_telefono", ""),
            formatear_moneda(s.get("valor", 0)),
            s.get("descripcion", "") or "",
            s["fecha"],
            dias_pend,
            "",
        ]
        for c, v in enumerate(fila, 1):
            cell = ws_resumen.cell(row=i, column=c, value=v)
            cell.font = money_font if c == 4 else cell_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if c in (1, 4, 6, 7) else "left")
            if fondo:
                cell.fill = fondo

    if todos_los_servicios:
        fila_total = len(todos_los_servicios) + 5
        ws_resumen.cell(row=fila_total, column=3, value="TOTAL GENERAL:").font = Font(bold=True, size=11)
        ws_resumen.cell(row=fila_total, column=4, value=formatear_moneda(total_general)).font = Font(bold=True, color="27ae60", size=12)

    for i, ancho in enumerate([8, 22, 15, 14, 35, 22, 14, 16], 1):
        ws_resumen.column_dimensions[get_column_letter(i)].width = ancho

    wb.save(ruta_destino)
    return ruta_destino
