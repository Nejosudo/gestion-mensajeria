import customtkinter as ctk
import tkinter.ttk as ttk
from core.config import COLORS, fmt_moneda, CTkToolTip
from database import database as db
from database.exportador import exportar_liquidaciones, exportar_servicios_pendientes
from CTkMessagebox import CTkMessagebox
from database.database import eliminar_liquidacion
from tkcalendar import DateEntry
import tkinter as tk
from tkinter import filedialog
from datetime import datetime


class TabFacturas(ctk.CTkFrame):
    def __init__(self, parent, **kwargs):
        super().__init__(parent, fg_color="transparent", **kwargs)
        self.pack(fill="both", expand=True)

        # Barra de filtros
        filtros_frame = ctk.CTkFrame(self, fg_color=COLORS["bg_card"], corner_radius=12)
        filtros_frame.pack(fill="x", padx=0, pady=(0, 10))

        ctk.CTkLabel(
            filtros_frame, text="🔍  Filtros Rápidos:",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=COLORS["text"]
        ).pack(side="left", padx=15, pady=12)
        self.filtro_var = ctk.StringVar(value="todo")

        # Configuración de Radios (SIN TEXTO NI ICONOS, SOLO TOOLTIP)
        opciones = [
            ("", "hoy", "Hoy"),
            ("", "semana", "Esta Semana"),
            ("", "mes", "Este Mes"),
            ("", "todo", "Todo el historial")
        ]

        for icono, valor, desc in opciones:
            rb = ctk.CTkRadioButton(
                filtros_frame, text=icono, variable=self.filtro_var, value=valor,
                fg_color=COLORS["accent"], hover_color=COLORS["accent_hover"],
                border_color=COLORS["border"], text_color=COLORS["text"],
                font=ctk.CTkFont(size=14), width=28,
                command=self._on_filter_changed
            )
            rb.pack(side="left", padx=2, pady=12)
            CTkToolTip(rb, desc)

        # Filtro por calendario (Condicional)
        self.rb_fecha = ctk.CTkRadioButton(
            filtros_frame, text="📆", variable=self.filtro_var, value="fecha",
            fg_color=COLORS["accent"], hover_color=COLORS["accent_hover"],
            border_color=COLORS["border"], text_color=COLORS["text"],
            font=ctk.CTkFont(size=14, weight="bold"), width=40,
            command=self._on_filter_changed
        )
        self.rb_fecha.pack(side="left", padx=(10, 5), pady=12)
        CTkToolTip(self.rb_fecha, "Filtrar por Rango de Fechas")

        # Contenedor de calendarios (OCULTO POR DEFECTO)
        self.cal_container = ctk.CTkFrame(filtros_frame, fg_color="transparent")
        
        # --- Estilo para calendario "Desde" ---
        self.lbl_desde = ctk.CTkLabel(self.cal_container, text="Desde:", font=ctk.CTkFont(size=12, weight="bold"), text_color=COLORS["text_muted"])
        self.lbl_desde.pack(side="left", padx=(5, 5))
        
        self.frame_desde = ctk.CTkFrame(self.cal_container, fg_color=COLORS["bg_input"], border_color=COLORS["border"], border_width=2, corner_radius=10, height=36)
        self.frame_desde.pack(side="left", padx=2)
        self.frame_desde.pack_propagate(False)

        self.cal_desde = DateEntry(
            self.frame_desde, width=12, background=COLORS["accent"],
            foreground=COLORS["text"], borderwidth=0, date_pattern='yyyy-mm-dd',
            locale='es_ES', font=("Segoe UI", 10),
            headersbackground=COLORS["accent"], headersforeground='white',
            selectbackground=COLORS["accent"], selectforeground='white',
            normalbackground=COLORS["bg_input"], normalforeground=COLORS["text"],
            weekendbackground=COLORS["bg_card"], weekendforeground=COLORS["text"],
            othermonthbackground=COLORS["table_row_2"], othermonthforeground=COLORS["text_muted"],
            relief="flat"
        )
        self.cal_desde.pack(padx=10, pady=0, fill="y", expand=True)
        
        # --- Estilo para calendario "Hasta" ---
        self.lbl_hasta = ctk.CTkLabel(self.cal_container, text="Hasta:", font=ctk.CTkFont(size=12, weight="bold"), text_color=COLORS["text_muted"])
        self.lbl_hasta.pack(side="left", padx=(10, 5))
        
        self.frame_hasta = ctk.CTkFrame(self.cal_container, fg_color=COLORS["bg_input"], border_color=COLORS["border"], border_width=2, corner_radius=10, height=36)
        self.frame_hasta.pack(side="left", padx=2)
        self.frame_hasta.pack_propagate(False)

        self.cal_hasta = DateEntry(
            self.frame_hasta, width=12, background=COLORS["accent"],
            foreground=COLORS["text"], borderwidth=0, date_pattern='yyyy-mm-dd',
            locale='es_ES', font=("Segoe UI", 10),
            headersbackground=COLORS["accent"], headersforeground='white',
            selectbackground=COLORS["accent"], selectforeground='white',
            normalbackground=COLORS["bg_input"], normalforeground=COLORS["text"],
            weekendbackground=COLORS["bg_card"], weekendforeground=COLORS["text"],
            othermonthbackground=COLORS["table_row_2"], othermonthforeground=COLORS["text_muted"],
            relief="flat"
        )
        self.cal_hasta.pack(padx=10, pady=0, fill="both", expand=True)
        
        self.cal_desde.bind("<<DateEntrySelected>>", lambda e: self._on_date_changed())
        self.cal_hasta.bind("<<DateEntrySelected>>", lambda e: self._on_date_changed())

        # Buscador por mensajero
        search_frame = ctk.CTkFrame(filtros_frame, fg_color=COLORS["bg_input"], corner_radius=8, height=34)
        search_frame.pack(side="right", padx=(0, 8), pady=12)
        search_frame.pack_propagate(False)
        ctk.CTkLabel(search_frame, text="🔍", font=ctk.CTkFont(size=13)).pack(side="left", padx=(8, 2))
        self.entry_buscar_liq = ctk.CTkEntry(
            search_frame, placeholder_text="Buscar mensajero...",
            fg_color="transparent", border_width=0,
            text_color=COLORS["text"], width=160
        )
        self.entry_buscar_liq.pack(side="left", fill="y", padx=(0, 6))
        self.entry_buscar_liq.bind("<KeyRelease>", lambda e: self._cargar_liquidaciones())


        # Tabla de liquidaciones
        tabla_liq_frame = ctk.CTkFrame(self, fg_color=COLORS["bg_card"], corner_radius=12)
        tabla_liq_frame.pack(fill="both", expand=True)

        self.tree_liquidaciones = ttk.Treeview(
            tabla_liq_frame,
            columns=("id", "mensajero", "fecha", "subtotal", "comision", "aseo", "base", "neto", "empresa", "num_servicios"),
            show="headings",
            style="Dark.Treeview",
            selectmode="browse"
        )
        for col, texto, ancho in [
            ("id", "ID", 40), ("mensajero", "Mensajero", 130),
            ("fecha", "Fecha", 150), ("subtotal", "Subtotal", 100),
            ("comision", "Comisión", 90), ("aseo", "Aseo", 60),
            ("base", "Base", 90), ("neto", "Neto Mens.", 110), 
            ("empresa", "Ganancia Emp.", 120), ("num_servicios", "N° Servicios", 120)
        ]:
            self.tree_liquidaciones.heading(col, text=texto)
            self.tree_liquidaciones.column(col, width=ancho, anchor="center")

        self.tree_liquidaciones.tag_configure("par", background=COLORS["table_row_2"])

        scrollbar_liq = ttk.Scrollbar(tabla_liq_frame, orient="vertical",
                                       command=self.tree_liquidaciones.yview)
        self.tree_liquidaciones.configure(yscrollcommand=scrollbar_liq.set)
        scrollbar_liq.pack(side="right", fill="y")
        self.tree_liquidaciones.pack(fill="both", expand=True, padx=5, pady=5)

        # Evento doble clic para mostrar tarjeta de liquidación
        self.tree_liquidaciones.bind("<Double-1>", self._abrir_tarjeta_liquidacion)

        # Resumen en la parte inferior
        self.resumen_frame = ctk.CTkFrame(self, fg_color=COLORS["bg_card"], corner_radius=12, height=60)
        self.resumen_frame.pack(fill="x", pady=(10, 0))
        self.resumen_frame.pack_propagate(False)

        # Botones de acción (Inferior Izquierda)
        self.btn_eliminar = ctk.CTkButton(
            self.resumen_frame, text="🗑️ Eliminar Liquidación", height=38, width=180,
            fg_color=COLORS["danger"], hover_color="#c0392b",
            text_color="white",
            font=ctk.CTkFont(size=12, weight="bold"),
            corner_radius=10,
            command=self._eliminar_liquidacion
        )
        self.btn_eliminar.pack(side="left", padx=15, pady=10)

        self.btn_exportar = ctk.CTkButton(
            self.resumen_frame, text="📥 Exportar Reporte", height=38, width=170,
            fg_color=COLORS["success"], hover_color="#219150",
            text_color="white",
            font=ctk.CTkFont(size=12, weight="bold"),
            corner_radius=10,
            command=self._exportar_excel
        )
        self.btn_exportar.pack(side="left", padx=(0, 15), pady=10)

        # Resumen (Derecha)
        self.lbl_resumen = ctk.CTkLabel(
            self.resumen_frame,
            text="Sin datos para mostrar",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=COLORS["text_muted"]
        )
        self.lbl_resumen.pack(side="right", padx=20)

        # Cargar liquidaciones al iniciar
        self.reload_data()

    def _abrir_tarjeta_liquidacion(self, event):
        if not self.tree_liquidaciones.winfo_exists():
            return
        seleccion = self.tree_liquidaciones.selection()
        if not seleccion:
            return
        item = self.tree_liquidaciones.item(seleccion[0])
        valores = item["values"]
        # Mapear columnas
        datos = {
            "ID": valores[0],
            "Mensajero": valores[1],
            "Fecha": valores[2],
            "Subtotal": valores[3],
            "Comisión": valores[4],
            "Aseo": valores[5],
            "Base": valores[6],
            "Neto Mens.": valores[7],
            "Ganancia Emp.": valores[8]
        }
        # Obtener servicios asociados a la liquidación
        try:
            id_liq = int(valores[0])
        except Exception:
            id_liq = None
        # Buscar datos de la liquidación en la base para obtener mensajero_id y fecha exacta
        filtro = self.filtro_var.get()
        liquidaciones = db.obtener_liquidaciones(filtro)
        liq = next((l for l in liquidaciones if str(l["id"]) == str(id_liq)), None)
        servicios = []
        if liq:
            servicios = db.obtener_servicios_por_liquidacion(liq["id"])
        # Centrar desde el principio
        self._mostrar_tarjeta_liquidacion(datos, servicios, parent=self)

    def _mostrar_tarjeta_liquidacion(self, datos, servicios, parent=None):
        if parent is None:
            parent = self
        ancho = 420
        alto = 600
        ventana = ctk.CTkToplevel(parent)
        ventana.title(f"Liquidación #{datos['ID']}")
        ventana.configure(fg_color=COLORS["bg_card"])
        ventana.transient(parent)
        # Centrar usando la ventana raíz real
        ventana.update()
        root = parent.winfo_toplevel()
        x = root.winfo_x() + (root.winfo_width() // 2) - (ancho // 2)
        y = root.winfo_y() + (root.winfo_height() // 2) - (alto // 2)
        ventana.geometry(f"{ancho}x{alto}+{x}+{y}")
        ventana.lift()
        ventana.after(100, ventana.grab_set)

        # Header estilo premium centrado
        header = ctk.CTkFrame(ventana, fg_color=COLORS["accent"], height=60, corner_radius=0)
        header.pack(fill="x")
        header.pack_propagate(False)
        ctk.CTkLabel(
            header, text=f"Detalle de Liquidación #{datos['ID']}",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color="#ffffff"
        ).pack(pady=16)

        # Contenedor principal
        main_frame = ctk.CTkFrame(ventana, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=25, pady=15)

        # Info Mensajero y Fecha
        ctk.CTkLabel(
            main_frame, text=f"👤 {datos['Mensajero']}",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=COLORS["text"]
        ).pack(anchor="w")
        ctk.CTkLabel(
            main_frame, text=f"📅 Fecha: {datos['Fecha']}",
            font=ctk.CTkFont(size=12),
            text_color=COLORS["text_muted"]
        ).pack(anchor="w", pady=(2, 10))

        # --- SECCIÓN SUPERIOR (Gris) ---
        f_superior = ctk.CTkFrame(main_frame, fg_color="#f8f9fa", corner_radius=12)
        f_superior.pack(fill="x", pady=10)

        # Servicios realizados
        f_srv = ctk.CTkFrame(f_superior, fg_color="transparent")
        f_srv.pack(fill="x", padx=15, pady=(15, 8))
        ctk.CTkLabel(f_srv, text="📦  Servicios liquidados", font=ctk.CTkFont(size=13), text_color=COLORS["text_muted"]).pack(side="left")
        ctk.CTkLabel(f_srv, text=str(len(servicios)), font=ctk.CTkFont(size=14, weight="bold"), text_color=COLORS["text"]).pack(side="right")

        # Subtotal generado
        f_sub = ctk.CTkFrame(f_superior, fg_color="transparent")
        f_sub.pack(fill="x", padx=15, pady=8)
        ctk.CTkLabel(f_sub, text="💰  Subtotal generado", font=ctk.CTkFont(size=13), text_color=COLORS["text_muted"]).pack(side="left")
        ctk.CTkLabel(f_sub, text=datos['Subtotal'], font=ctk.CTkFont(size=14, weight="bold"), text_color=COLORS["text"]).pack(side="right")

        # Pago a Mensajero (Rojo y Negativo)
        f_pago = ctk.CTkFrame(f_superior, fg_color="transparent")
        f_pago.pack(fill="x", padx=15, pady=(8, 15))
        ctk.CTkLabel(f_pago, text="🏍️   Pago a Mensajero", font=ctk.CTkFont(size=13), text_color=COLORS["text_muted"]).pack(side="left")
        ctk.CTkLabel(f_pago, text=f"- {datos['Neto Mens.']}", font=ctk.CTkFont(size=14, weight="bold"), text_color=COLORS["danger"]).pack(side="right")

        # Separador
        ctk.CTkFrame(main_frame, height=1, fg_color=COLORS["border"]).pack(fill="x", pady=15)

        # --- SECCIÓN INFERIOR ---
        # Ganancia Empresa
        f_ganancia = ctk.CTkFrame(main_frame, fg_color="transparent")
        f_ganancia.pack(fill="x", pady=4)
        ctk.CTkLabel(f_ganancia, text="🏢  GANANCIA EMPRESA:", font=ctk.CTkFont(size=13, weight="bold"), text_color=COLORS["text"]).pack(side="left")
        ctk.CTkLabel(f_ganancia, text=datos['Comisión'], font=ctk.CTkFont(size=16, weight="bold"), text_color=COLORS["success"]).pack(side="right")

        # Aseo
        f_aseo = ctk.CTkFrame(main_frame, fg_color="transparent")
        f_aseo.pack(fill="x", pady=4)
        ctk.CTkLabel(f_aseo, text="🖌️  ASEO:", font=ctk.CTkFont(size=13, weight="bold"), text_color=COLORS["text"]).pack(side="left")
        ctk.CTkLabel(f_aseo, text=datos['Aseo'], font=ctk.CTkFont(size=16, weight="bold"), text_color=COLORS["success"]).pack(side="right")

        # Base a Devolver
        f_base = ctk.CTkFrame(main_frame, fg_color="transparent")
        f_base.pack(fill="x", pady=(4, 15))
        ctk.CTkLabel(f_base, text="🏠  BASE A DEVOLVER:", font=ctk.CTkFont(size=13, weight="bold"), text_color=COLORS["text"]).pack(side="left")
        ctk.CTkLabel(f_base, text=datos['Base'], font=ctk.CTkFont(size=16, weight="bold"), text_color=COLORS["warning"]).pack(side="right")

        # Separador para los servicios si existen
        if servicios:
            ctk.CTkLabel(main_frame, text="Detalle de servicios:", font=ctk.CTkFont(size=14, weight="bold"), text_color=COLORS["accent"]).pack(anchor="w", pady=(5, 5))
            svc_frame = ctk.CTkScrollableFrame(main_frame, fg_color=COLORS["bg_input"], corner_radius=8, height=150)
            svc_frame.pack(fill="x", pady=(0, 10))
            for s in servicios:
                desc = s.get("descripcion", "") or ""
                texto_principal = f"🚴 ID: {s['id']}  |  💰 {fmt_moneda(s['valor'])}"
                fila = ctk.CTkFrame(svc_frame, fg_color="transparent")
                fila.pack(fill="x", padx=4, pady=2)
                ctk.CTkLabel(fila, text=texto_principal, font=ctk.CTkFont(size=11, weight="bold"), text_color=COLORS["text"], anchor="w").pack(anchor="w")
                if desc:
                    ctk.CTkLabel(fila, text=f"   📝 {desc}", font=ctk.CTkFont(size=10), text_color=COLORS["text_muted"], anchor="w").pack(anchor="w")
                ctk.CTkFrame(svc_frame, height=1, fg_color=COLORS["border"]).pack(fill="x", padx=4)
        else:
            ctk.CTkLabel(main_frame, text="No se encontraron servicios asociados.", font=ctk.CTkFont(size=12), text_color=COLORS["text_muted"]).pack(anchor="w", pady=2)

    def _cargar_liquidaciones(self):
        """Recarga la tabla de liquidaciones según el filtro seleccionado y la búsqueda."""
        for item in self.tree_liquidaciones.get_children():
            self.tree_liquidaciones.delete(item)

        filtro = self.filtro_var.get()
        if filtro == "fecha":
            f_inicio = self.cal_desde.get_date().strftime("%Y-%m-%d")
            f_fin = self.cal_hasta.get_date().strftime("%Y-%m-%d")
            filtro = f"{f_inicio}..{f_fin}"
            
        liquidaciones = db.obtener_liquidaciones(filtro)

        # Aplicar filtro de búsqueda por mensajero
        busqueda = self.entry_buscar_liq.get().strip().lower() if hasattr(self, 'entry_buscar_liq') else ""
        if busqueda:
            liquidaciones = [l for l in liquidaciones if busqueda in (l.get("mensajero_nombre") or "").lower()]

        total_neto = 0
        total_comision = 0

        for i, liq in enumerate(liquidaciones):
            tags = ("par",) if i % 2 == 1 else ()
            ganancia_empresa = liq["comision_empresa"] + liq["descuento_aseo"]
            num_servicios = liq.get("num_servicios", 0)
            self.tree_liquidaciones.insert("", "end", values=(
                liq["id"],
                liq.get("mensajero_nombre", ""),
                liq["fecha"],
                fmt_moneda(liq["subtotal_servicios"]),
                fmt_moneda(liq["comision_empresa"]),
                fmt_moneda(liq["descuento_aseo"]),
                fmt_moneda(liq.get("base_prestada", 0)),
                fmt_moneda(liq["neto_mensajero"]),
                fmt_moneda(ganancia_empresa),
                num_servicios
            ), tags=tags)
            total_neto += liq["neto_mensajero"]
            total_comision += liq["comision_empresa"]

        total_aseo = sum(l["descuento_aseo"] for l in liquidaciones)
        total_empresa = total_comision + total_aseo

        if liquidaciones:
            self.lbl_resumen.configure(
                text=f"📊  {len(liquidaciones)} liquidaciones  |  "
                     f"💰 Total Neto: {fmt_moneda(total_neto)}  |  "
                     f"🏢 Total Comisiones: {fmt_moneda(total_comision)}",
                text_color=COLORS["accent"]
            )
        else:
            self.lbl_resumen.configure(
                text="Sin liquidaciones para el filtro seleccionado.",
                text_color=COLORS["text_muted"]
            )
    def reload_data(self):
        self._cargar_liquidaciones()

    def _on_filter_changed(self):
        self._actualizar_estado_calendarios()
        self._cargar_liquidaciones()

    def _actualizar_estado_calendarios(self):
        es_fecha = self.filtro_var.get() == "fecha"
        if es_fecha:
            self.cal_container.pack(side="left", padx=5)
        else:
            self.cal_container.pack_forget()

    def _exportar_excel(self):
        """Exporta las liquidaciones visibles a un archivo Excel
        e incluye una hoja extra con los servicios pendientes actuales."""
        filtro = self.filtro_var.get()
        if filtro == "fecha":
            f_inicio = self.cal_desde.get_date().strftime("%Y-%m-%d")
            f_fin = self.cal_hasta.get_date().strftime("%Y-%m-%d")
            filtro = f"{f_inicio}..{f_fin}"
        datos = db.obtener_liquidaciones(filtro)

        if not datos:
            CTkMessagebox(title="ℹ️ Sin datos", message="No hay liquidaciones para exportar.",
                          icon="info", option_1="OK")
            return

        # Preguntar ubicación al usuario
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        ruta_sugerida = f"Liquidaciones_{timestamp}.xlsx"
        
        ruta_destino = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            initialfile=ruta_sugerida,
            title="Seleccionar dónde guardar el reporte"
        )
        
        if not ruta_destino:
            return

        try:
            # 1. Exportar liquidaciones
            ruta = exportar_liquidaciones(datos, ruta_destino=ruta_destino)

            # 2. Agregar hoja de servicios pendientes al mismo archivo
            self._agregar_hoja_pendientes(ruta)

            # Preguntar si desea abrir la ubicación o ver el archivo
            msg = CTkMessagebox(
                title="✅ Exportación Exitosa",
                message=f"Archivo generado correctamente.\n\nIncluye las liquidaciones y una hoja extra con los servicios pendientes de todos los mensajeros.",
                icon="check", option_1="Abrir archivo", option_2="OK"
            )
            
            if msg.get() == "Abrir archivo":
                try:
                    import subprocess, platform
                    if platform.system() == 'Darwin':
                        subprocess.call(('open', ruta))
                    elif platform.system() == 'Windows':
                        os.startfile(ruta)
                    else:
                        subprocess.call(('xdg-open', ruta))
                except Exception:
                    pass
        except Exception as e:
            CTkMessagebox(
                title="❌ Error",
                message=f"No se pudo exportar:\n{str(e)}",
                icon="cancel", option_1="OK"
            )

    def _agregar_hoja_pendientes(self, ruta_xlsx: str):
        """Abre el Excel ya generado y le agrega una hoja con servicios pendientes."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime

        wb = openpyxl.load_workbook(ruta_xlsx)

        # Estilos
        header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill  = PatternFill(start_color="16213e", end_color="16213e", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        title_font   = Font(name="Calibri", bold=True, size=13, color="16213e")
        sub_font     = Font(name="Calibri", italic=True, size=9, color="666666")
        cell_font    = Font(name="Calibri", size=10)
        money_font   = Font(name="Calibri", size=10, color="27ae60")
        warn_fill    = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
        alt_fill     = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        ws = wb.create_sheet(title="Servicios Pendientes")

        mensajeros = db.obtener_mensajeros()
        todos = []
        for m in mensajeros:
            for s in db.obtener_servicios_pendientes(m["id"]):
                todos.append({**s, "mensajero_nombre": m["nombre"], "mensajero_telefono": m["telefono"]})

        # Título
        ws.merge_cells("A1:H1")
        ws["A1"].value = "SERVICIOS PENDIENTES DE LIQUIDAR — RESPALDO"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:H2")
        ws["A2"].value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}   |   Total servicios pendientes: {len(todos)}"
        ws["A2"].font = sub_font
        ws["A2"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A3:H3")
        ws["A3"].value = "⚠️  Estos servicios aún NO han sido liquidados. Incluidos como respaldo de seguridad."
        ws["A3"].font = Font(name="Calibri", bold=True, size=10, color="856404")
        ws["A3"].fill = warn_fill
        ws["A3"].alignment = Alignment(horizontal="center")

        headers = ["ID", "Mensajero", "Teléfono", "Valor", "Descripción / Cliente", "Fecha y Hora", "Días pendiente", "Base Mensajero"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
        ws.row_dimensions[4].height = 22

        hoy = datetime.now().date()
        total_general = 0.0

        if not todos:
            ws.merge_cells("A5:H5")
            ws["A5"].value = "No hay servicios pendientes en este momento."
            ws["A5"].font = Font(name="Calibri", italic=True, size=11, color="888888")
            ws["A5"].alignment = Alignment(horizontal="center")
        else:
            for i, s in enumerate(todos, 5):
                try:
                    dias = (hoy - datetime.strptime(s["fecha"], "%Y-%m-%d %H:%M:%S").date()).days
                except Exception:
                    dias = "?"
                total_general += s.get("valor", 0)
                fondo = alt_fill if (i - 5) % 2 == 1 else None
                fila = [
                    s.get("id", ""),
                    s.get("mensajero_nombre", ""),
                    s.get("mensajero_telefono", ""),
                    f"${s.get('valor', 0):,.0f}".replace(",", "."),
                    s.get("descripcion", "") or "",
                    s["fecha"],
                    dias,
                    "",
                ]
                for c, v in enumerate(fila, 1):
                    cell = ws.cell(row=i, column=c, value=v)
                    cell.font = money_font if c == 4 else cell_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center" if c in (1, 4, 6, 7) else "left")
                    if fondo:
                        cell.fill = fondo

            fila_total = len(todos) + 5
            ws.cell(row=fila_total, column=3, value="TOTAL GENERAL:").font = Font(bold=True, size=11)
            ws.cell(row=fila_total, column=4, value=f"${total_general:,.0f}".replace(",", ".")).font = Font(bold=True, color="27ae60", size=12)

        for i, ancho in enumerate([8, 22, 15, 14, 35, 22, 14, 16], 1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

        wb.save(ruta_xlsx)
    
    def _on_date_changed(self):
        """Al cambiar fecha en el calendario, activa el radio de fecha y recarga."""
        self.filtro_var.set("fecha")
        self._on_filter_changed()

    def _eliminar_liquidacion(self):
        """Elimina la liquidación seleccionada tras confirmar con contraseña."""
        seleccion = self.tree_liquidaciones.selection()
        if not seleccion:
            CTkMessagebox(
                title="⚠️ Sin selección",
                message="Selecciona una liquidación de la tabla para eliminarla.",
                icon="warning", option_1="OK"
            )
            return

        valores = self.tree_liquidaciones.item(seleccion[0], "values")
        id_liq   = valores[0]
        mensajero = valores[1]
        fecha    = valores[2]

        # Confirmación previa
        confirm = CTkMessagebox(
            title="🗑️ Eliminar Liquidación",
            message=(
                f"¿Estás seguro de eliminar la liquidación #{id_liq}?\n\n"
                f"👤 Mensajero: {mensajero}\n"
                f"📅 Fecha: {fecha}\n\n"
                f"⚠️ Los servicios de esta liquidación volverán a quedar PENDIENTES."
            ),
            icon="question", option_1="Cancelar", option_2="Eliminar"
        )
        if confirm.get() != "Eliminar":
            return

        # Pedir contraseña
        self._solicitar_password(
            titulo="🔒 Contraseña requerida",
            mensaje="Ingresa la contraseña para eliminar esta liquidación:",
            callback=lambda: self._ejecutar_eliminacion(int(id_liq))
        )

    def _ejecutar_eliminacion(self, id_liq: int):
        """Realiza la eliminación efectiva de la liquidación."""
        try:
            eliminar_liquidacion(id_liq)
            self._cargar_liquidaciones()
            # Notificar a tab_gestion si existe (para actualizar pendientes)
            CTkMessagebox(
                title="✅ Eliminado",
                message=f"La liquidación #{id_liq} fue eliminada correctamente.\nLos servicios volvieron a estado pendiente.",
                icon="check", option_1="OK"
            )
        except Exception as e:
            CTkMessagebox(
                title="❌ Error",
                message=f"No se pudo eliminar la liquidación:\n{e}",
                icon="cancel", option_1="OK"
            )

    def _solicitar_password(self, titulo: str, mensaje: str, callback):
        """Muestra un modal centrado pidiendo la contraseña de administrador."""
        import customtkinter as ctk
        root = self.winfo_toplevel()
        modal = ctk.CTkToplevel(root)
        modal.title(titulo)
        modal.geometry("320x180")
        modal.configure(fg_color=COLORS["bg_card"])
        modal.transient(root)
        modal.resizable(False, False)

        modal.update_idletasks()
        x = root.winfo_x() + (root.winfo_width() // 2) - 160
        y = root.winfo_y() + (root.winfo_height() // 2) - 90
        modal.geometry(f"+{x}+{y}")
        modal.after(100, modal.grab_set)

        ctk.CTkLabel(modal, text=mensaje, font=ctk.CTkFont(size=12),
                     wraplength=280).pack(pady=(20, 8), padx=20)
        entry_pass = ctk.CTkEntry(modal, width=200, show="*")
        entry_pass.pack(pady=5)
        entry_pass.focus_set()

        def verificar(event=None):
            if entry_pass.get() == db.get_app_password():
                modal.destroy()
                callback()
            else:
                CTkMessagebox(title="Error", message="Contraseña incorrecta.", icon="cancel")
                entry_pass.delete(0, "end")
                entry_pass.focus_set()

        entry_pass.bind("<Return>", verificar)
        ctk.CTkButton(
            modal, text="Verificar", fg_color=COLORS["accent"],
            hover_color=COLORS["accent_hover"], command=verificar
        ).pack(pady=15)
